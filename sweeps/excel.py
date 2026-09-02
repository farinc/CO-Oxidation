"""The main output of the sweeps is an .xlsx workbook with a Parameters/Axes/Grid details
and per-step sheets (kMC trajectories, mean-field time series, ME-MKM joint
adsorbate-count distributions + eigenvalues), and, if coexistence ran, a Coexistence
summary sheet plus one ME-MKM sheet per crossing.

Sheet names are capped at S{step_index}_{kind} / C{crossing_index}_MEMKM
N-D grid step is traced back to parameter values via the Grid sheet rather 
than encoded in the sheet name itself.

The ME-MKM sheets hold only the adsorbate-count ("joint distribution") grids and
the eigenvalue list. The plotting side still gets full per-microstate fidelity (it 
takes the SweepResult, whose CoexistenceCrossing.arrays.
"""

from dataclasses import fields, replace

import numpy as np
import openpyxl
import pandas as pd

from co_oxidation.observables import array_items, scalar_items

from .axes import _INT_AXES, AxisSpec, GridStep, SweepGrid
from .params import RunConfig
from .results import (CoexistenceCrossing, KMCTrajectoryResult, MeanFieldModelResult,
                      MemkmStepResult, StepResult, SweepResult)

_MEANFIELD_SHEET_KIND = {"mf": "MFMKM", "ea": "BW"}
_MEANFIELD_COLS = ["tag", "t", "theta_co", "theta_o", "theta_empty", "r_ads_co",
                   "r_des_co", "r_ads_o", "r_oxi", "r_des_o"]


# --- write ---------------------------------------------------------------------------

def write_workbook(result: SweepResult, path) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        _write_parameters_sheet(writer, result.grid.base, result.grid.axes)
        _write_axes_sheet(writer, result.grid.axes)
        _write_grid_sheet(writer, result.grid)

        _write_observables_sheet(writer, result)

        for step_result in result.steps:
            idx = step_result.step.index
            if step_result.kmc is not None:
                _write_kmc_sheet(writer, f"S{idx}_kMC", step_result.kmc)
            if step_result.meanfield is not None:
                _write_meanfield_sheets(writer, idx, step_result.meanfield)
            if step_result.memkm is not None:
                _write_memkm_sheet(writer, f"S{idx}_MEMKM", step_result.memkm)
            arrays = array_items(step_result.observables or {})
            if arrays:
                _write_observable_arrays_sheet(writer, f"S{idx}_OBS", arrays)

        if result.coexistence:
            _write_coexistence_sheet(writer, result.coexistence)
            for i, crossing in enumerate(result.coexistence):
                _write_memkm_crossing_sheet(writer, f"C{i}_MEMKM", crossing)


def _write_parameters_sheet(writer, base: RunConfig, axes: list[AxisSpec]):
    swept_names = {a.name for a in axes}
    rows = [{"name": f.name, "value": getattr(base, f.name), "swept": f.name in swept_names}
            for f in fields(base)]
    pd.DataFrame(rows).to_excel(writer, sheet_name="Parameters", index=False)


def _write_axes_sheet(writer, axes: list[AxisSpec]):
    rows = [{"name": a.name, "values": ",".join(str(v) for v in a.values)} for a in axes]
    pd.DataFrame(rows, columns=["name", "values"]).to_excel(
        writer, sheet_name="Axes", index=False)


def _write_grid_sheet(writer, grid: SweepGrid):
    axis_names = [a.name for a in grid.axes]
    rows = [{"step_index": s.index, **{n: s.axis_values[n] for n in axis_names}}
            for s in grid.steps]
    pd.DataFrame(rows, columns=["step_index", *axis_names]).to_excel(
        writer, sheet_name="Grid", index=False)


def _write_observables_sheet(writer, result: SweepResult):
    """One row per grid step, one column per scalar --observable value, so a sweep's
    derived quantities can be read (and plotted) straight off a single sheet next to
    the Grid sheet that says what each step's parameters were.

    Columns are the union over steps, in first-seen order: an observable that only
    fires on some steps leaves blanks rather than shifting anyone else's column. Array
    values do not belong in a one-row-per-step table and go to S{i}_OBS instead."""
    rows, columns = [], []
    for step_result in result.steps:
        values = scalar_items(step_result.observables or {})
        if not values:
            continue
        rows.append({"step_index": step_result.step.index, **values})
        for key in values:
            if key not in columns:
                columns.append(key)
    if not rows:
        return
    pd.DataFrame(rows, columns=["step_index", *columns]).to_excel(
        writer, sheet_name="Observables", index=False)


def _write_observable_arrays_sheet(writer, sheet_name, arrays: dict):
    """The array-valued --observable results for one step, one column each, NaN-padded
    to the longest. Multidimensional arrays are flattened in C order (the observable
    knows its own shape; the sheet is a transport format, not a structure), and a
    complex array is split into {key}_re / {key}_im since a cell holds one real."""
    columns = {}
    for key, value in arrays.items():
        flat = np.asarray(value).ravel()
        if np.iscomplexobj(flat):
            columns[f"{key}_re"] = flat.real
            columns[f"{key}_im"] = flat.imag
        else:
            columns[key] = flat.astype(float)
    max_len = max(len(c) for c in columns.values())
    padded = {k: np.concatenate([v, np.full(max_len - len(v), np.nan)])
              for k, v in columns.items()}
    pd.DataFrame(padded).to_excel(writer, sheet_name=sheet_name, index=False)


def _write_kmc_sheet(writer, sheet_name, trajectories: list[KMCTrajectoryResult]):
    summary = pd.DataFrame([{
        "tag": t.tag, "trajectory_index": t.index, "seed": t.seed,
        "steady_empty": t.steady_empty, "steady_co": t.steady_co,
        "steady_o": t.steady_o, "t_final": t.t_final, "steps": t.steps,
        "stuck": t.stuck,
    } for t in trajectories])
    summary.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)

    max_len = max((len(t.times) for t in trajectories), default=0)
    series = {}
    for t in trajectories:
        prefix = f"{t.tag}{t.index}"
        pad = max_len - len(t.times)
        for col, arr in (("t", t.times), ("cov_empty", t.cov_empty),
                        ("cov_co", t.cov_co), ("cov_o", t.cov_o)):
            series[f"{col}_{prefix}"] = np.concatenate([arr, np.full(pad, np.nan)])
    startrow = len(summary) + 2
    pd.DataFrame(series).to_excel(writer, sheet_name=sheet_name, index=False,
                                  startrow=startrow)


def _write_meanfield_sheets(writer, step_index, results: list[MeanFieldModelResult]):
    by_model: dict[str, list[MeanFieldModelResult]] = {}
    for r in results:
        by_model.setdefault(r.model, []).append(r)
    for model, kind in _MEANFIELD_SHEET_KIND.items():
        frames = [pd.DataFrame({
            "tag": r.tag, "t": r.t, "theta_co": r.theta_co, "theta_o": r.theta_o,
            "theta_empty": r.theta_empty, "r_ads_co": r.r_ads_co,
            "r_des_co": r.r_des_co, "r_ads_o": r.r_ads_o, "r_oxi": r.r_oxi,
            "r_des_o": r.r_des_o,
        }) for r in by_model.get(model, [])]
        df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(
            columns=_MEANFIELD_COLS)
        df.to_excel(writer, sheet_name=f"S{step_index}_{kind}", index=False)


def _write_grid_block(writer, sheet_name, title, grid, row):
    """Writes a one-cell title at 0-indexed row `row`, then `grid` (shape (l+1, l+1),
    [N_CO, N_O]) transposed to row-index N_O / column-index N_CO immediately below.
    Returns the next free 0-indexed row, leaving one blank row after the table."""
    ws = writer.sheets[sheet_name]
    ws.cell(row=row + 1, column=1, value=title)
    l = grid.shape[0] - 1
    df = pd.DataFrame(grid.T, index=[f"N_O={i}" for i in range(l + 1)],
                      columns=[f"N_CO={j}" for j in range(l + 1)])
    df.index.name = "N_O\\N_CO"
    df.to_excel(writer, sheet_name=sheet_name, startrow=row + 1, index=True)
    return row + 2 + len(df) + 1


def _write_memkm_sheet(writer, sheet_name, memkm: MemkmStepResult):
    summary = pd.DataFrame([{
        "n_sites": memkm.n_sites, "theta_empty": memkm.theta_empty,
        "theta_co": memkm.theta_co, "theta_o": memkm.theta_o, "p_a": memkm.p_a,
        "empty_a": memkm.empty_a, "co_a": memkm.co_a, "o_a": memkm.o_a,
        "empty_b": memkm.empty_b, "co_b": memkm.co_b, "o_b": memkm.o_b,
    }])
    summary.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
    row = len(summary) + 2

    eigvals = np.asarray(memkm.eigvals)
    eig_df = pd.DataFrame({"index": np.arange(len(eigvals)),
                           "eigval_real": eigvals.real, "eigval_imag": eigvals.imag})
    eig_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=row)
    row += len(eig_df) + 2

    phi = memkm.cov_phi if memkm.cov_phi is not None else np.full_like(memkm.cov_pop, np.nan)
    row = _write_grid_block(writer, sheet_name, "Population P(N_CO,N_O)", memkm.cov_pop, row)
    row = _write_grid_block(writer, sheet_name, "psi_R_2 (sum)", memkm.cov_r2, row)
    row = _write_grid_block(writer, sheet_name, "psi_L_2 (class mean)", phi, row)
    row = _write_grid_block(writer, sheet_name, "Degeneracy", memkm.cov_deg, row)


def _write_coexistence_sheet(writer, crossings: list[CoexistenceCrossing]):
    rows = [{**c.outer_axis_values, "bisection_axis": c.bisection_axis,
            "value_star": c.value_star, **c.row} for c in crossings]
    pd.DataFrame(rows).to_excel(writer, sheet_name="Coexistence", index=False)


def _write_memkm_crossing_sheet(writer, sheet_name, crossing: CoexistenceCrossing):
    arrays, row = crossing.arrays, crossing.row
    # report()'s row already carries the basin-conditioned coverages and P_A; the
    # overall means are their basin-weighted average, P_A*x_A + P_B*x_B.
    theta_co = row["P_A"] * row["co_A"] + row["P_B"] * row["co_B"]
    theta_o = row["P_A"] * row["o_A"] + row["P_B"] * row["o_B"]
    theta_empty = row["P_A"] * row["empty_A"] + row["P_B"] * row["empty_B"]
    memkm = MemkmStepResult(
        eigvals=arrays["eigvals"], cov_pop=arrays["cov_pop"], cov_r2=arrays["cov_r2"],
        cov_phi=arrays["cov_phi"], cov_deg=arrays["cov_deg"], theta_empty=theta_empty,
        theta_co=theta_co, theta_o=theta_o, n_sites=arrays["n_sites"], p_a=row["P_A"],
        empty_a=row["empty_A"], co_a=row["co_A"], o_a=row["o_A"],
        empty_b=row["empty_B"], co_b=row["co_B"], o_b=row["o_B"])
    _write_memkm_sheet(writer, sheet_name, memkm)


# --- read (standalone replot) ---------------------------------------------------------

def _iter_blocks(ws):
    """Yield lists of non-empty rows (each a tuple of cell values, trimmed to its own
    used width), splitting on fully-blank rows."""
    block = []
    for row in ws.iter_rows(values_only=True):
        if all(v is None for v in row):
            if block:
                yield block
                block = []
            continue
        last = max(i for i, v in enumerate(row) if v is not None)
        block.append(row[:last + 1])
    if block:
        yield block


def _block_to_df(block):
    header = block[0]
    rows = [r + (None,) * (len(header) - len(r)) for r in block[1:]]
    return pd.DataFrame(rows, columns=header)


def _read_kmc_sheet(wb, sheet_name) -> list[KMCTrajectoryResult]:
    blocks = list(_iter_blocks(wb[sheet_name]))
    summary = _block_to_df(blocks[0])
    series = _block_to_df(blocks[1]) if len(blocks) > 1 else pd.DataFrame()
    out = []
    for _, r in summary.iterrows():
        prefix = f"{r['tag']}{int(r['trajectory_index'])}"

        def col(name):
            key = f"{name}_{prefix}"
            arr = series[key].to_numpy(dtype=float) if key in series.columns else np.array([])
            return arr[~np.isnan(arr)] if arr.size else arr

        out.append(KMCTrajectoryResult(
            tag=r["tag"], index=int(r["trajectory_index"]), seed=int(r["seed"]),
            times=col("t"), cov_empty=col("cov_empty"), cov_co=col("cov_co"),
            cov_o=col("cov_o"), steady_empty=float(r["steady_empty"]),
            steady_co=float(r["steady_co"]), steady_o=float(r["steady_o"]),
            t_final=float(r["t_final"]), steps=int(r["steps"]), stuck=bool(r["stuck"])))
    return out


def _read_meanfield_sheet(wb, sheet_name, model) -> list[MeanFieldModelResult]:
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    df = pd.DataFrame(rows[1:], columns=rows[0])
    out = []
    for tag, sub in df.groupby("tag"):
        sub = sub.sort_values("t")
        out.append(MeanFieldModelResult(
            model=model, tag=tag, t=sub["t"].to_numpy(float),
            theta_co=sub["theta_co"].to_numpy(float), theta_o=sub["theta_o"].to_numpy(float),
            theta_empty=sub["theta_empty"].to_numpy(float),
            r_ads_co=sub["r_ads_co"].to_numpy(float), r_des_co=sub["r_des_co"].to_numpy(float),
            r_ads_o=sub["r_ads_o"].to_numpy(float), r_oxi=sub["r_oxi"].to_numpy(float),
            r_des_o=sub["r_des_o"].to_numpy(float)))
    return out


def _read_memkm_sheet(wb, sheet_name) -> MemkmStepResult:
    blocks = list(_iter_blocks(wb[sheet_name]))
    summary = _block_to_df(blocks[0]).iloc[0]
    eig_df = _block_to_df(blocks[1])
    eigvals = eig_df["eigval_real"].to_numpy(float) + 1j * eig_df["eigval_imag"].to_numpy(float)

    def grid_from(block):
        df = _block_to_df(block[1:])   # block[0] is the lone title row
        return df.iloc[:, 1:].to_numpy(float).T   # back to [N_CO, N_O]

    cov_pop = grid_from(blocks[2])
    cov_r2 = grid_from(blocks[3])
    cov_phi = grid_from(blocks[4])
    cov_deg = grid_from(blocks[5])
    if np.isnan(cov_phi).all():
        cov_phi = None
    return MemkmStepResult(
        eigvals=eigvals, cov_pop=cov_pop, cov_r2=cov_r2, cov_phi=cov_phi, cov_deg=cov_deg,
        theta_empty=float(summary["theta_empty"]), theta_co=float(summary["theta_co"]),
        theta_o=float(summary["theta_o"]), n_sites=int(summary["n_sites"]),
        p_a=float(summary["p_a"]), empty_a=float(summary["empty_a"]),
        co_a=float(summary["co_a"]), o_a=float(summary["o_a"]),
        empty_b=float(summary["empty_b"]), co_b=float(summary["co_b"]),
        o_b=float(summary["o_b"]))


def _read_observables(path, wb, sheetnames) -> dict[int, dict]:
    """Rebuild each step's observable dict from the Observables sheet (scalars) and its
    S{i}_OBS sheet (arrays). Arrays come back flattened -- the sheet never recorded
    their shape -- and a complex array comes back through its _re/_im column pair."""
    by_step: dict[int, dict] = {}
    if "Observables" in sheetnames:
        df = pd.read_excel(path, sheet_name="Observables")
        for _, row in df.iterrows():
            values = {k: v for k, v in row.items()
                      if k != "step_index" and not pd.isna(v)}
            by_step.setdefault(int(row["step_index"]), {}).update(values)
    for name in sheetnames:
        if not (name.startswith("S") and name.endswith("_OBS")):
            continue
        idx = int(name[1:-len("_OBS")])
        df = pd.read_excel(path, sheet_name=name)
        arrays = {}
        for col in df.columns:
            trimmed = df[col].dropna().to_numpy()
            if col.endswith("_re"):
                key = col[:-3]
                arrays[key] = trimmed + 1j * np.zeros_like(trimmed)
            elif col.endswith("_im"):
                key = col[:-3]
                if key in arrays:
                    arrays[key] = arrays[key].real + 1j * trimmed
            else:
                arrays[col] = trimmed
        by_step.setdefault(idx, {}).update(arrays)
    return by_step


def read_workbook(path) -> SweepResult:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheetnames = set(wb.sheetnames)

    params_df = pd.read_excel(path, sheet_name="Parameters")
    base_kwargs = dict(zip(params_df["name"], params_df["value"]))
    for name in _INT_AXES:
        if name in base_kwargs and base_kwargs[name] is not None:
            base_kwargs[name] = int(base_kwargs[name])
    base = RunConfig(**base_kwargs)

    axes_df = pd.read_excel(path, sheet_name="Axes")
    axes = []
    for _, r in axes_df.iterrows():
        cast = int if r["name"] in _INT_AXES else float
        axes.append(AxisSpec(name=r["name"],
                             values=tuple(cast(v) for v in str(r["values"]).split(","))))
    axis_names = [a.name for a in axes]

    grid_df = pd.read_excel(path, sheet_name="Grid")
    steps = []
    for _, r in grid_df.iterrows():
        axis_values = {n: (int(r[n]) if n in _INT_AXES else float(r[n])) for n in axis_names}
        steps.append(GridStep(index=int(r["step_index"]), axis_values=axis_values,
                              config=replace(base, **axis_values)))
    grid = SweepGrid(base=base, axes=axes, steps=steps)

    observables_by_step = _read_observables(path, wb, sheetnames)

    step_results = []
    for step in steps:
        kmc_name = f"S{step.index}_kMC"
        kmc = _read_kmc_sheet(wb, kmc_name) if kmc_name in sheetnames else None

        mf_results = []
        for model, kind in _MEANFIELD_SHEET_KIND.items():
            name = f"S{step.index}_{kind}"
            if name in sheetnames:
                mf_results.extend(_read_meanfield_sheet(wb, name, model))

        memkm_name = f"S{step.index}_MEMKM"
        memkm = _read_memkm_sheet(wb, memkm_name) if memkm_name in sheetnames else None

        step_results.append(StepResult(step=step, kmc=kmc, meanfield=mf_results or None,
                                       memkm=memkm,
                                       observables=observables_by_step.get(step.index, {})))

    coexistence = None
    if "Coexistence" in sheetnames:
        coex_df = pd.read_excel(path, sheet_name="Coexistence")
        meta_cols = {"bisection_axis", "value_star"}
        coexistence = []
        for i, r in coex_df.iterrows():
            bisection_axis = r["bisection_axis"]
            outer_axis_values = {n: r[n] for n in axis_names
                                 if n != bisection_axis and n in coex_df.columns}
            row = {k: v for k, v in r.items()
                  if k not in meta_cols and k not in outer_axis_values}
            memkm_name = f"C{i}_MEMKM"
            memkm = _read_memkm_sheet(wb, memkm_name) if memkm_name in sheetnames else None
            arrays = {} if memkm is None else {
                "eigvals": memkm.eigvals, "cov_pop": memkm.cov_pop, "cov_r2": memkm.cov_r2,
                "cov_phi": memkm.cov_phi, "cov_deg": memkm.cov_deg, "n_sites": memkm.n_sites,
            }
            coexistence.append(CoexistenceCrossing(
                outer_axis_values=outer_axis_values, bisection_axis=bisection_axis,
                value_star=float(r["value_star"]), row=row, arrays=arrays,
                scan_values=np.array([]), scan_log_ratios=np.array([])))

    return SweepResult(grid=grid, steps=step_results, coexistence=coexistence)
